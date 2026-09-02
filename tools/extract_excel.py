# -*- coding: utf-8 -*-
"""
스노우베어베이커리 상품정보.xlsx -> data/products.json + images/{full,thumb}

엑셀 구조 (시트 '상품리스트(이미지)')
  1행: 그룹헤더(이미지 / 코드 / 기본정보 / 소분작업)
  2행: 컬럼명
  3행~: 데이터 (내용 없는 행은 건너뜀)
  이미지: A열=소분전, B열=소분후 (셀 앵커 기준으로 행에 귀속)

사용:
  python tools/extract_excel.py "<엑셀경로>" --out .
  python tools/extract_excel.py "<엑셀경로>" --out . --merge

  --merge 를 주면 기존 products.json 을 읽어 판매코드/상품명이 매칭되는 건은
  갱신하고 없는 건만 추가한다. 엑셀에 없는 기존 상품은 지우지 않는다.
"""
import argparse
import hashlib
import io
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone, timedelta

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image

KST = timezone(timedelta(hours=9))

SHEET_PRODUCTS = "상품리스트(이미지)"
SHEET_MATERIALS = "부자재"

HEADER_ROW = 2
DATA_START_ROW = 3

# 엑셀 컬럼명 -> JSON 키
COLUMN_MAP = {
    "준비주체2": "prep_owner",
    "대분류5": "main_category",
    "푸드레인 판매코드": "foodrain_code",
    "실제판매코드": "sale_code",
    "ERP상품명": "erp_name",
    "스노우베이커리 상품코드2": "sb_code",
    "스노우베이커리 명칭": "sb_name",
    "매입사명": "vendor",
    "보관방법": "storage",
    "유통기한": "shelf_life",
    "제품사이즈": "product_size",
    "규격": "spec",
    "소분규격": "portion_spec",
    "입수": "pack_qty",
    "분류": "work_type",
    "개별 포장여부": "individual_pack",
    "표시사항 부착여부": "label_attach",
    "소분포장방식": "pack_method",
    "포장지": "bag",
    "트레이": "tray",
    "라벨 (투명)": "label",
    "추가작업": "extra_work",
    "시식": "tasting",
    "가열방법": "heating",
}

IMAGE_COLUMNS = {0: "images_before", 1: "images_after"}  # 0-based: A열, B열

THUMB_MAX = 200
THUMB_QUALITY = 82

FORBIDDEN_IN_FILENAME = re.compile(r'[\\/:*?"<>|\s]+')


def clean(v):
    """셀 값 정규화. 앞뒤 공백/개행 제거, 빈 문자열은 None."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.replace("\r\n", "\n")
        s = re.sub(r"[ \t]+\n", "\n", s)
        s = re.sub(r"\n[ \t]+", "\n", s)
        s = s.strip()
        return s or None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return v


def slugify(text, fallback):
    """이미지 파일명용 슬러그. 한글은 유지하고 파일명 금지문자만 치환."""
    if not text:
        return fallback
    s = FORBIDDEN_IN_FILENAME.sub("_", str(text).strip())
    s = re.sub(r"_+", "_", s).strip("_")
    return s[:40] or fallback


def read_products(ws):
    """헤더행에서 컬럼 매핑을 만들고 데이터 행을 파싱한다."""
    col_to_key = {}
    unmapped = []
    for c in range(1, ws.max_column + 1):
        name = clean(ws.cell(HEADER_ROW, c).value)
        if not name:
            continue
        if name in COLUMN_MAP:
            col_to_key[c] = COLUMN_MAP[name]
        elif name not in ("이미지", "소분전", "소분후"):
            unmapped.append(get_column_letter(c) + ":" + str(name))
    if unmapped:
        print("  [경고] JSON 매핑이 없는 컬럼: " + ", ".join(unmapped), file=sys.stderr)

    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        rec = {}
        for c, key in col_to_key.items():
            val = clean(ws.cell(r, c).value)
            if val is not None:
                rec[key] = val
        if not rec:
            continue
        rec["_row"] = r
        rows.append(rec)
    return rows


def extract_images(ws, rows, out_dir):
    """셀 앵커 기준으로 이미지를 행에 귀속. 원본 PNG + WebP 썸네일 저장."""
    full_dir = os.path.join(out_dir, "images", "full")
    thumb_dir = os.path.join(out_dir, "images", "thumb")
    os.makedirs(full_dir, exist_ok=True)
    os.makedirs(thumb_dir, exist_ok=True)

    by_cell = defaultdict(list)
    for im in ws._images:
        anchor = im.anchor._from
        if anchor.col not in IMAGE_COLUMNS:
            continue
        by_cell[(anchor.row + 1, anchor.col)].append(im._data())

    row_index = {rec["_row"]: rec for rec in rows}
    seen_hash = {}
    saved = 0

    for (excel_row, col), blobs in sorted(by_cell.items()):
        rec = row_index.get(excel_row)
        if rec is None:
            print("  [경고] %d행 이미지가 상품 행과 매칭되지 않아 건너뜀" % excel_row,
                  file=sys.stderr)
            continue
        key = IMAGE_COLUMNS[col]
        suffix = "before" if col == 0 else "after"
        base = rec["id"] + "_" + slugify(rec.get("erp_name"), rec["id"])

        for i, blob in enumerate(blobs, 1):
            digest = hashlib.md5(blob).hexdigest()
            if digest in seen_hash:
                rec.setdefault(key, []).append(seen_hash[digest])
                continue

            name = "%s_%s%d.png" % (base, suffix, i)
            with open(os.path.join(full_dir, name), "wb") as f:
                f.write(blob)

            img = Image.open(io.BytesIO(blob))
            if img.mode not in ("RGB", "RGBA"):
                img = img.convert("RGBA")
            img.thumbnail((THUMB_MAX, THUMB_MAX), Image.LANCZOS)
            img.save(os.path.join(thumb_dir, name[:-4] + ".webp"),
                     "WEBP", quality=THUMB_QUALITY)

            seen_hash[digest] = name
            rec.setdefault(key, []).append(name)
            saved += 1

    return saved


def read_materials(wb):
    if SHEET_MATERIALS not in wb.sheetnames:
        return []
    ws = wb[SHEET_MATERIALS]
    out = []
    for r in range(2, ws.max_row + 1):
        kind = clean(ws.cell(r, 1).value)
        name = clean(ws.cell(r, 2).value)
        link = clean(ws.cell(r, 3).value)
        if not name:
            continue
        out.append({"kind": kind, "name": name, "link": link})
    return out


def next_id(records):
    nums = [int(r["id"][1:]) for r in records
            if isinstance(r.get("id"), str) and re.fullmatch(r"P\d+", r["id"])]
    return "P%03d" % ((max(nums) + 1) if nums else 1)


def merge(rows, products_path):
    with open(products_path, encoding="utf-8") as f:
        existing = json.load(f)
    old = existing.get("products", [])

    def match_key(r):
        return (r.get("sale_code") or r.get("foodrain_code")
                or r.get("erp_name") or r.get("id"))

    index = {match_key(r): r for r in old}
    added = updated = 0
    for rec in rows:
        target = index.get(match_key(rec))
        if target is not None:
            rec["id"] = target["id"]
            rec["active"] = target.get("active", True)
            # 이번 엑셀에서 이미지가 안 나왔으면 기존 이미지를 유지한다
            for key in ("images_before", "images_after"):
                if not rec.get(key):
                    rec[key] = target.get(key, [])
            target.clear()
            target.update(rec)
            updated += 1
        else:
            rec["id"] = next_id(old)
            old.append(rec)
            added += 1
    print("  병합: %d건 갱신, %d건 추가" % (updated, added))
    return old


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel")
    ap.add_argument("--out", default=".")
    ap.add_argument("--merge", action="store_true",
                    help="기존 products.json 과 병합 (엑셀에 없는 상품은 보존)")
    args = ap.parse_args()

    out_dir = os.path.abspath(args.out)
    data_dir = os.path.join(out_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    print("엑셀 읽는 중: " + args.excel)
    wb = openpyxl.load_workbook(args.excel, data_only=True)
    if SHEET_PRODUCTS not in wb.sheetnames:
        sys.exit("시트 '%s' 를 찾을 수 없습니다. 있는 시트: %s"
                 % (SHEET_PRODUCTS, wb.sheetnames))
    ws = wb[SHEET_PRODUCTS]

    rows = read_products(ws)
    print("  상품 %d건" % len(rows))

    for i, rec in enumerate(rows, 1):
        rec["id"] = "P%03d" % i
        rec["active"] = True

    saved = extract_images(ws, rows, out_dir)
    print("  이미지 %d장 저장 (원본 PNG + WebP 썸네일)" % saved)

    now = datetime.now(KST).isoformat(timespec="seconds")
    for rec in rows:
        rec.pop("_row", None)
        rec.setdefault("images_before", [])
        rec.setdefault("images_after", [])
        rec["updated_at"] = now

    products_path = os.path.join(data_dir, "products.json")
    if args.merge and os.path.exists(products_path):
        rows = merge(rows, products_path)

    payload = {
        "updated_at": now,
        "source": os.path.basename(args.excel),
        "count": len(rows),
        "products": rows,
    }
    with open(products_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)
    print("  -> data/products.json")

    materials = read_materials(wb)
    if materials:
        with open(os.path.join(data_dir, "materials.json"), "w", encoding="utf-8") as f:
            json.dump({"updated_at": now, "materials": materials}, f,
                      ensure_ascii=False, indent=1)
        print("  -> data/materials.json (%d건)" % len(materials))


if __name__ == "__main__":
    main()
